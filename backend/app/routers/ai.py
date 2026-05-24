import json
import os
import tempfile

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from app.core.database import get_db
from app.core.config import settings
from app.core.security import get_current_user_id
from app.models.models import Question, ImportLog, Category
from app.schemas.admin import QuestionCreate

router = APIRouter(prefix="/admin/ai", tags=["AI识别"])


async def _call_ai(prompt: str, text_content: str) -> str:
    try:
        from openai import AsyncOpenAI
        client = AsyncOpenAI(api_key=settings.AI_API_KEY, base_url=settings.AI_API_URL or None)
        response = await client.chat.completions.create(
            model=settings.AI_MODEL_NAME or "gpt-4o-mini",
            messages=[
                {"role": "system", "content": prompt},
                {"role": "user", "content": text_content},
            ],
            temperature=0.1,
        )
        return response.choices[0].message.content or ""
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"AI调用失败: {str(e)}")


def _extract_json(text: str) -> list:
    text = text.strip()
    if text.startswith("```"):
        lines = text.split("\n")
        text = "\n".join(lines[1:-1] if lines[-1].strip() == "```" else lines[1:])
    try:
        result = json.loads(text)
        if isinstance(result, list):
            return result
        if isinstance(result, dict) and "questions" in result:
            return result["questions"]
        return [result]
    except json.JSONDecodeError:
        start = text.find("[")
        end = text.rfind("]") + 1
        if start >= 0 and end > start:
            try:
                return json.loads(text[start:end])
            except json.JSONDecodeError:
                pass
        raise HTTPException(status_code=500, detail="AI返回内容无法解析为JSON")


async def _read_file(file: UploadFile) -> str:
    suffix = os.path.splitext(file.filename or "")[1].lower()
    content = await file.read()

    if suffix in (".txt", ".md"):
        return content.decode("utf-8", errors="ignore")

    if suffix == ".pdf":
        try:
            from PyPDF2 import PdfReader
            import io
            reader = PdfReader(io.BytesIO(content))
            return "\n".join(page.extract_text() or "" for page in reader.pages)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"PDF解析失败: {str(e)}")

    if suffix in (".docx", ".doc"):
        try:
            from docx import Document
            import io
            doc = Document(io.BytesIO(content))
            return "\n".join(p.text for p in doc.paragraphs)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"DOCX解析失败: {str(e)}")

    if suffix in (".xlsx", ".xls"):
        try:
            from openpyxl import load_workbook
            import io
            wb = load_workbook(io.BytesIO(content))
            rows = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    rows.append(" | ".join(str(c) for c in row if c))
            return "\n".join(rows)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Excel解析失败: {str(e)}")

    return content.decode("utf-8", errors="ignore")


@router.post("/recognize/interview")
async def recognize_interview(
    file: UploadFile = File(...),
    category_id: int | None = Form(None),
    db: AsyncSession = Depends(get_db),
    user_id: int = Depends(get_current_user_id),
):
    text = await _read_file(file)

    prompt = """你是一个面试题识别专家。请从以下文本中提取所有面试题，返回JSON数组。
每个元素包含：
- content: 题目内容（必填）
- oral_answer: 口述答案（如有）
- ref_answer: 参考答案（如有）
- difficulty: 难度（初阶/中阶/高阶）
- tags: 标签数组
- source: 来源

只返回JSON，不要其他文字。"""

    ai_result = await _call_ai(prompt, text)
    items = _extract_json(ai_result)

    if category_id:
        for item in items:
            item.setdefault("category_id", category_id)
    for item in items:
        item.setdefault("question_type", "interview")

    return {"code": 200, "count": len(items), "items": items}


@router.post("/recognize/written")
async def recognize_written(
    file: UploadFile = File(...),
    category_id: int | None = Form(None),
    db: AsyncSession = Depends(get_db),
    user_id: int = Depends(get_current_user_id),
):
    text = await _read_file(file)

    prompt = """你是一个笔试题识别专家。请从以下文本中提取所有笔试题，返回JSON数组。
每个元素包含：
- content: 题目内容（必填）
- ref_answer: 参考答案（如有）
- difficulty: 难度（初阶/中阶/高阶）
- tags: 标签数组
- source: 来源

只返回JSON，不要其他文字。"""

    ai_result = await _call_ai(prompt, text)
    items = _extract_json(ai_result)

    if category_id:
        for item in items:
            item.setdefault("category_id", category_id)
    for item in items:
        item.setdefault("question_type", "written")

    return {"code": 200, "count": len(items), "items": items}


@router.post("/recognize/project")
async def recognize_project(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user_id: int = Depends(get_current_user_id),
):
    text = await _read_file(file)

    prompt = """你是一个简历/项目分析专家。请从以下文本中提取所有项目经历，返回JSON数组。
每个元素包含：
- name: 项目名称（必填）
- description: 项目描述
- tech_stack: 技术栈数组
- duties: 项目职责
- highlights: 项目亮点

只返回JSON，不要其他文字。"""

    ai_result = await _call_ai(prompt, text)
    items = _extract_json(ai_result)

    return {"code": 200, "count": len(items), "items": items}


@router.post("/import/save")
async def save_ai_import(
    questions: list[QuestionCreate],
    db: AsyncSession = Depends(get_db),
    user_id: int = Depends(get_current_user_id),
):
    created = []
    for item in questions:
        question = Question(**item.model_dump())
        db.add(question)
        created.append(question)

    import_log = ImportLog(
        user_id=user_id,
        file_name="ai_import",
        file_type="ai",
        parse_type="ai",
        total_count=len(created),
        status="success",
    )
    db.add(import_log)

    await db.flush()
    return {"code": 200, "message": f"成功入库 {len(created)} 条", "count": len(created)}
